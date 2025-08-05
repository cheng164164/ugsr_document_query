import azure.functions as func
import logging
import os
import json
from openai import AzureOpenAI
import requests
import re
from datetime import datetime, timedelta
from azure.storage.blob import BlobServiceClient, BlobClient, generate_blob_sas, BlobSasPermissions
import json
import io
import openpyxl
import pandas as pd


# Load environment variables
AZURE_BLOB_CONN_STRING = os.getenv("AZURE_BLOB_CONN_STRING")
AZURE_SEARCH_ENDPOINT = os.getenv("AZURE_SEARCH_ENDPOINT")
AZURE_SEARCH_KEY = os.getenv("AZURE_SEARCH_KEY")
AZURE_OPENAI_ENDPOINT = os.getenv("AZURE_OPENAI_ENDPOINT")
AZURE_OPENAI_API_KEY = os.getenv("AZURE_OPENAI_API_KEY")
AZURE_OPENAI_DEPLOYMENT = os.getenv("AZURE_OPENAI_DEPLOYMENT")
AZURE_OPENAI_EMBEDDING_DEPLOYMENT = "text-embedding-3-small"


# Define mutli indexes names to search
index_names = ["business_index",
               "ugsr_index"
              ]

# Metadata files in blob storage for each index
metadata_files = {
    "business_index": {'container_name': "north-america-business-documents-metadata", 'file_name': "auto_extraction/business_metadata_new.csv"},
    "ugsr_index": {'container_name': "undergroound-engineering-document-metadata", 'file_name': "auto_extraction/ugsr_metadata_new.csv"}
}

# SharePoint URLs for each index
share_point_urls = {
    "business_index": {'name': 'Business Documents North America', "url": "https://globalkomatsu.sharepoint.com/sites/NAGMUSGR00243/SitePages/PublishedDocuments.aspx"},
    "ugsr_index": {'name': 'UGSR Engineering Documents', "url": "https://globalkomatsu.sharepoint.com/sites/NAGMUSGR00221/engres_joy/PPED/JGUEngDocs?viewpath=%2Fsites%2FNAGMUSGR00221%2Fengres%5Fjoy%2FPPED%2FJGUEngDocs"}
}

# Feature On/Off flags
debug_mode = True   # Set to True to enable debug prints
custom_ranking = True   # Set to True to enable custom ranking (vector + keyword); False to use Azure default ranking
dynamic_filtering = False   # Set to True to enable dynamic metadata filtering based on query keywords
keywords_matching = True   # Set to True to enable keyword matching check and warning    
metadata_search = True    # Set to True to enable metadata-only search for relevant queries
use_prev_context = True   # Set to True to enable the feature that uses previous queries as context
hide_ref_relevance = True # Set to True to hide relevance explanation in the reference section

app = func.FunctionApp(http_auth_level=func.AuthLevel.FUNCTION)


def clean_query_for_llm(raw_query, route_keywords={"metadata", "content", "contents"}):
    """
    Removes standalone routing keywords from the query if they are not part of a sentence.
    Argument:  raw_query  -  can be single query or query history (sperated by '|')
    """
    try:
        # Normalize and split query into clauses using punctuation
        parts = re.split(r'[,.!?;|\n]+', raw_query)
        cleaned = []

        for part in parts:
            stripped = part.strip()
            # If this part is exactly one of the routing keywords, skip it
            if stripped.lower() in route_keywords:
                continue
            # If it's empty or only a keyword, skip
            if not stripped or all(w.lower() in route_keywords for w in stripped.split()):
                continue
            cleaned.append(stripped)

        if '|' in raw_query:   # If input is query history (sperated by '|')
            return '| '.join(cleaned).strip()
        
        else:   #  If inout is single query
            return '. '.join(cleaned).strip()
    
    except Exception as e:
        logging.error(f"Error cleaning query: {e}")
        return raw_query


def filter_relevant_history(current_query, query_history, answer_history):
    """
    Filters relevant user-bot turns from chat history based on the current query.
    Returns a list of relevant exchanges (dicts with 'user' and 'bot').
    """
    if not current_query.strip():
        return ""

    queries = [q.strip() for q in query_history.split("|") if q.strip()]
    answers = [a.strip() for a in answer_history.split("|") if a.strip()]

    history_turns = []
    max_len = max(len(queries), len(answers))

    for i in range(max_len):
        user = queries[i] if i < len(queries) else None
        bot = answers[i] if i < len(answers) else None
        turn_number = i+1
        turn_lines = [f"Turn {turn_number}:"]
        if user and bot:
            turn_lines.append(f"User: {user}\nBot: {bot}")
        elif user:
            turn_lines.append(f"User: {user}")
        elif bot:
            turn_lines.append(f"Bot: {bot}")

        history_turns.append("\n".join(turn_lines))

    if not history_turns:
        return ""

    client = AzureOpenAI(
        azure_endpoint=AZURE_OPENAI_ENDPOINT,
        api_key=AZURE_OPENAI_API_KEY,
        api_version="2024-12-01-preview"
    )

    transcript = "\n\n".join(history_turns)

    prompt = (
        "You are an assistant that filters past chat history to keep only what is useful for understanding the current question.\n"
        "Each turn may include a user question, a bot answer, or both.\n"
        "Your task is to return only the relevant items from the history that help clarify or add context to the current question.\n"
        "Ignore unrelated entries.\n\n"
        "If the CURRENT QUESTION appears vague (e.g., contains 'it' or 'that'), lacking of context and reference. In this case, "
        "prioritize the most recent turns (e.g., Turn 5 is newer than Turn 1) that might clarify those references, and ignore older, unrelated entries.\n"
        f"CURRENT QUESTION:\n{current_query}\n\n"
        f"PAST HISTORY:\n{transcript}"
    )

    response = client.chat.completions.create(
        model=AZURE_OPENAI_DEPLOYMENT,
        messages=[{"role": "user", "content": prompt}]
    )

    return response.choices[0].message.content.strip()


def rewrite_query_with_history(current_query, relevant_history_text):
    """
    Rewrites the current query using relevant chat history for clarity.
    """
    if not relevant_history_text or not current_query.strip():
        return current_query

    client = AzureOpenAI(
        azure_endpoint=AZURE_OPENAI_ENDPOINT,
        api_key=AZURE_OPENAI_API_KEY,
        api_version="2024-12-01-preview"
    )

    system_prompt = (
        "You are a smart query cleaner and rewriter. A user is asking a question that may refer to earlier exchanges.\n"
        "You are given the current question and a relevant excerpt from prior user and assistant turns.\n"
        "Your job is to:\n"
        "- Clarify the current question by resolving any vague terms (e.g., 'it', 'this', 'that', 'these', 'those', 'they', 'the one') using the history.\n"
        "- Keep only what is necessary to make the current query clear.\n"
        "- Do not list the history or answer the question.\n"
        "Return only the rewritten version of the current question."
    )

    user_prompt = (
        f"RELEVANT HISTORY:\n{relevant_history_text}\n\n"
        f"CURRENT QUESTION:\n{current_query}"
    )

    response = client.chat.completions.create(
        model=AZURE_OPENAI_DEPLOYMENT,
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt}
        ]
    )

    return response.choices[0].message.content.strip()


def llm_search_query_optimizer(query, rewrited_query, use_previous_context):
    system_prompt = f"You are a search query optimizer. \
        Read the user question and extract the key words to generate search queries to improve semantic matching for vector search.\
        Make sure your rewritten query includes all important original terms (do not drop any).\
        but also add relevant synonyms and related descriptions.\
        Remove stopwords.\
        Return a single line of plain text that expands the original query, not a list. Do not return JSON, quotes, or explanations"


    client = AzureOpenAI(
        azure_endpoint=AZURE_OPENAI_ENDPOINT,
        api_key=AZURE_OPENAI_API_KEY,
        api_version="2024-12-01-preview"
    )

    if rewrited_query and use_previous_context:
        query_input = f"Context: {rewrited_query}\nOriginal query: {query}"
    else:
        query_input = f"Original query: {query}"
    

    response = client.chat.completions.create(
                        model="o3-mini",
                        messages = [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": query_input}
        ]  
                                )

    optimized_query = response.choices[0].message.content.strip().lower()
    return optimized_query


def get_query_embedding(query):
    client = AzureOpenAI(
        azure_endpoint=AZURE_OPENAI_ENDPOINT,
        api_key=AZURE_OPENAI_API_KEY,
        api_version="2024-12-01-preview"
    )

    embedding_response = client.embeddings.create(
        model=AZURE_OPENAI_EMBEDDING_DEPLOYMENT,
        input=query
    )
    return embedding_response.data[0].embedding

def extract_keywords(query, optimized_query, debug=False):
    original_keywords = set(re.findall(r"\b\w+\b", query.lower()))
    optimized_keywords = set(re.findall(r"\b\w+\b", optimized_query.lower()))
    stopwords = {"how", "can", "the", "is", "what", "my", "i", "you", "your", "to", "a", "and", "in", "of", "for", 
                 "on", "with", "that", "this", "it", "as", "an", "by", "at", "from", "be", "are", "or", "if", "but", 
                 "not", "all", "any", "so", "do", "does", "did", "have", "has", "had", "documents", "document", "documentation", "documentations"}
    intersection_keywords = (original_keywords.intersection(optimized_keywords)) - stopwords
    if debug:
        print(f"Original query: {query.lower()}")
        print(f"Optimized query: {optimized_query.lower()}")
        print(f"Original query keywords: {original_keywords}")
        print(f"Optimized query keywords: {optimized_keywords}")
        print(f"intersection keywords: {intersection_keywords}")
    return intersection_keywords


def normalize_plural_keywords(keywords):
    """
    Normalize plural keywords to match singular variations like 'forms' -> 'form', 'processes' -> 'process'.
    """
    normalized = set(keywords)
    for kw in keywords:
        if kw.endswith("es"):
            normalized.add(kw[:-2])  # handles e.g., "processes" -> "process"
            normalized.add(kw[:-1])  # also handles e.g., "titles" -> "title"
        elif kw.endswith("s"):
            normalized.add(kw[:-1])  # handles e.g., "forms" -> "form"
    return list(normalized)


def fetch_metadata_from_index(headers, index_name, keywords, fields, sample_size=500):
    """
    Fetch a small number of documents from each index to collect known metadata field values.
    """
    metadata_docs = []
    normalized_keywords = normalize_plural_keywords(keywords)  # handle plural
    url = f"{AZURE_SEARCH_ENDPOINT}/indexes/{index_name}/docs/search?api-version=2024-07-01"
    payload = {
        "search": " ".join(normalized_keywords),
        "top": sample_size,
        "queryType": "simple",         # enables keyword search with implicit OR logic
        "searchFields": ",".join(fields),  # restricts scope
        "searchMode": "any"
    }
    response = requests.post(url, headers=headers, json=payload)
    if response.status_code == 200:
        metadata_docs.extend(response.json().get("value", []))
    return metadata_docs, normalized_keywords


def collect_field_values(docs, fields):
    """
    Collect unique values for each metadata field across all documents.
    Returns a dict like { "doc_type": {"form", "procedure"}, ... }
    """
    values_by_field = {field: set() for field in fields}
    for doc in docs:
        for field in fields:
            val = doc.get(field, "").lower()
            if val:
                values_by_field[field].add(val)
    return values_by_field


def generate_or_filter_from_keywords(keywords, field_value_dict):
    """
    Build an OData filter clause using OR logic where keywords match metadata field values.
    """
    clauses = []
    for field, values in field_value_dict.items():
        for val in values:
            for kw in keywords:
                if kw in val:
                    clauses.append(f"{field} eq '{val}'")
                    break
    return " or ".join(clauses)


def generate_field_based_filter(headers, payload, index_name, keywords, filter_fields):
    """
    Use keyword-based OR filter if any metadata fields match and apply to payload.
    """
    metadata_docs, normalized_keywords = fetch_metadata_from_index(headers, index_name, keywords, filter_fields, sample_size=500)
    field_values = collect_field_values(metadata_docs, filter_fields)
    or_data_filter = generate_or_filter_from_keywords(normalized_keywords, field_values)
    if or_data_filter:
        payload["filter"] = or_data_filter


def cosine_similarity(vec1, vec2):
    dot = sum(a * b for a, b in zip(vec1, vec2))
    norm1 = sum(a * a for a in vec1) ** 0.5
    norm2 = sum(b * b for b in vec2) ** 0.5
    return dot / (norm1 * norm2 + 1e-8)


def generate_blob_sas_url(connection_string, container_name, blob_name):
    blob_service_client = BlobServiceClient.from_connection_string(connection_string)
    blob_client = blob_service_client.get_blob_client(container=container_name, blob=blob_name)

    account_key = blob_service_client.credential.account_key
    sas_token = generate_blob_sas(
        account_name=blob_client.account_name,
        container_name=container_name,
        blob_name=blob_name,
        account_key=account_key,
        permission=BlobSasPermissions(read=True),
        expiry=datetime.utcnow() + timedelta(hours=1)
    )

    return f"{blob_client.url}?{sas_token}"


def should_use_metadata_search(query):
    client = AzureOpenAI(
        azure_endpoint=AZURE_OPENAI_ENDPOINT,
        api_key=AZURE_OPENAI_API_KEY,
        api_version="2024-12-01-preview"
    )

    messages = [
        {
            "role": "system",
            "content": (
                "You are a router. Your job is to decide whether a query should be answered by:\n"
                "- metadata: if it asks for listings, like filenames, titles, document types, categories, release version, revision date, release data of the documents in the resources etc.\n"
                "- semantic: if it needs detailed answers from document content.\n"
                "If the query is ambiguous or general, prefer semantic.\n"
                "If the query contains words like 'metadata', choose metadata. If the query contains words like 'content', choose semantic.\n"
                "Only reply with one word: 'metadata' or 'semantic'."
            )
        },
        {
            "role": "user",
            "content": f"Query: {query}"
        }
    ]

    response = client.chat.completions.create(
        model=AZURE_OPENAI_DEPLOYMENT,
        messages=messages
    )

    decision = response.choices[0].message.content.strip().lower()
    print(f"* Routing decision: {decision} *\n")
    return decision == "metadata"


def llm_context_guard_check(query, context_text, client, deployment=AZURE_OPENAI_DEPLOYMENT):
    """
    Uses LLM to confirm whether the provided context actually answers the user's query.
    Returns a tuple (is_valid, explanation)
    """
    system_msg = {
        "role": "system",
        "content": (
            "You are a validation agent. Your job is to decide if the provided CONTEXT truly answers the USER QUESTION.\n"
            "Be strict. If the context uses different terms, systems, or services than the question, reply 'no'.\n"
            "Do not infer answers. Only consider exact term matches.\n"
            "Your reply must start with 'yes' or 'no'. Then give a brief reason why."
        )
    }
    user_msg = {
        "role": "user",
        "content": (
            f"USER QUESTION: {query}\n\n"
            f"CONTEXT: {context_text}\n\n"
            f"Does the context fully and specifically answer the question based on term and system alignment?"
        )
    }
    response = client.chat.completions.create(
        model=deployment,
        messages=[system_msg, user_msg]
    )
    answer = response.choices[0].message.content.strip().lower()
    is_valid = answer.startswith("yes")

    # Second: determine if the question is completely irrelevant
    is_completely_irrelevant = False
    if not is_valid:
        irrelevance_messages =[{
                "role": "user",
                "content": (
                    f"Determine whether the following question is completely unrelated to the provided context.\n"
                    f"Only perform this check if the question cannot be answered using the context.\n"
                    f"If the question is at least somewhat related to the context, respond with RELEVANT.\n"
                    f"If it is completely off-topic or unrelated, respond with IRRELEVANT.\n"
                    f"QUESTION: {query}\n"
                    f"CONTEXT:{context_text}\n"
                    f"INSTRUCTIONS: \n"
                    f" - If at least partial of keywords or terms match between the user query and provded context, then consider it as RELEVANT. \n" 
                    f" - Reply with one word only: RELEVANT or IRRELEVANT."
                )
            }
        ]        

    irrelevance_response = client.chat.completions.create(
        model=deployment,
        messages=irrelevance_messages
    )

    relevance_tag = irrelevance_response.choices[0].message.content.strip().upper()
    is_completely_irrelevant = relevance_tag == "IRRELEVANT"

    return is_valid, answer, is_completely_irrelevant


def metadata_table_by_index(index_names):
    metadata_by_index = {}
    for index_name in index_names:
        if index_name in metadata_files:
            container_name = metadata_files[index_name]['container_name']
            file_name = metadata_files[index_name]['file_name']
            try:
                blob = BlobClient.from_connection_string(AZURE_BLOB_CONN_STRING, container_name, file_name)
                data = blob.download_blob().readall()
            except Exception as e:
                logging.error(f"[Azure] Blob fetch error: {type(e).__name__} - {str(e)}")
                metadata_by_index[index_name] = [{"error": f"Azure Blob error: {type(e).__name__} - {str(e)}"}]
                continue
            
            try:
                if file_name.endswith(".csv"):
                    df = pd.read_csv(io.BytesIO(data))
                elif file_name.endswith(".xlsx") or file_name.endswith(".xls"):
                    df = pd.read_excel(io.BytesIO(data), engine="openpyxl")
                else:
                    raise ValueError("Unsupported file type")

                df = df.fillna("").astype(str)
                metadata_by_index[index_name] = df.to_dict(orient="records")
            except Exception as e:
                logging.error(f"[Pandas] Excel parse error: {type(e).__name__} - {str(e)}")
                metadata_by_index[index_name] = [{"error": f"Pandas Excel error: {type(e).__name__} - {str(e)}"}]
        
    return metadata_by_index


'''
def metadata_table_by_index(index_names):
    metadata_by_index = {}
    BLOB_URLS_BY_INDEX = {}
    for index_name in index_names:
        if index_name in metadata_files:
            container_name = metadata_files[index_name]['container_name']
            file_name = metadata_files[index_name]['file_name']
            BLOB_URLS_BY_INDEX[index_name] = generate_blob_sas_url(AZURE_BLOB_CONN_STRING, container_name, file_name)

    for index_name, blob_url in BLOB_URLS_BY_INDEX.items():
        response = requests.get(blob_url)
        if b"<?xml" in response.content or b"AccessDenied" in response.content or b"<Error>" in response.content:
            metadata_by_index[index_name] = [{"error": "Failed to fetch metadata. Access is restricted or URL is invalid."}]
            print("error: ", "Failed to fetch metadata. Access is restricted or URL is invalid.")
            continue
        
        # df = pd.read_excel(io.BytesIO(response.content))
        # df = df.fillna("").astype(str)
        # metadata_by_index[index_name] = df.to_dict(orient="records")
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content), data_only=True)
        sheet = wb.active
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        records = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            record = {headers[i]: str(cell).strip() if cell is not None else "" for i, cell in enumerate(row)}
            records.append(record)
        metadata_by_index[index_name] = records
        
    return metadata_by_index
'''

def summarize_full_metadata(query, relevant_history_text, metadata_by_index):
    client = AzureOpenAI(
        azure_endpoint=AZURE_OPENAI_ENDPOINT,
        api_key=AZURE_OPENAI_API_KEY,
        api_version="2024-12-01-preview"
    )

    index_contexts = []
    for index_name, docs in metadata_by_index.items():
        context_lines = [f"\nIndex: {index_name}"]
        for doc in docs:
            row_desc = "; ".join(f"{k}: {v}" for k, v in doc.items())
            context_lines.append(f"- {row_desc}")
        index_contexts.append("\n".join(context_lines))
    full_context = "\n\n".join(index_contexts)

    history_prefix = (
        f"RELEVANT CHAT HISTORY:\n{relevant_history_text}\n\n"
        if relevant_history_text else ""
    )

    prompt = (
        f"You are an assistant that summarizes document metadata.\n"
        f"{history_prefix}"
        f"USER QUERY:\n{query}\n\n"
        f"DOCUMENT METADATA:\n{full_context}\n\n"
        f"Instructions:\n"
        f"- Use history only if it helps clarify the current query.\n"
        f"-Only answer the current query. Do not answer or repeat previous questions.\n"
        f"-If the query mentions a specific index, only summarize that index. Otherwise, summarize all indexes.\n"
        f"-list as many relevant documents as possible that match the user query.\n"
        f"-If you are not sure about the answer or nothing relevant is found, say 'Sorry, I cannot help with it. Please try looking it up on above share point links'.\n"
        f"-Use clear bullet points or sections."
    )
    
    completion = client.chat.completions.create(
        model=AZURE_OPENAI_DEPLOYMENT,
        messages=[{"role": "user", "content": prompt}]
    )
    summary = completion.choices[0].message.content.strip()

    # Append SharePoint links as references
    reference_links = "\n\n\nYou can look up metadata information from the SharePoint.\n\n"+ "**SharePoint Links:**\n"

    for index, values in share_point_urls.items():
        url = values.get("url", "")
        if url:
            reference_links += f"- {values.get('name', 'Unknown')} : {url}\n\n"

    return  f"**Answer:**\n\n{reference_links} \n\n\n\n {summary}"         


def multi_index_search_documents(query, rewrited_query, index_names, vector_weight=0.6, top_k=6, 
                                 dynamic_filtering=True, keywords_matching=True, custom_ranking=True, 
                                 use_previous_context = True, debug=False):
    """
    Performs hybrid search across multiple indexes.
    Args:
        query (str): The search query.
        index_names (list): List of Azure Search index names.
        vector_weight (float): Weight for vector similarity in final ranking [0.0 - 1.0].
        top_k (int): Number of top documents to return per index.
        dynamic_filtering (bool): If True, apply dynamic metadata filtering based on query keywords.
        keywords_matching (bool): If True, check if keywords are present in retrieved documents and warn if missing.
        custom_ranking (bool): If True, manually calculate final score using vector + keyword; otherwise use Azure's ranking.
        use_previous_context (bool): If True, rewrite current query using previous queries as context.
        debug (bool): If True, print debug output
    Returns:
        list of documents with relevance scores and index tags.
    """

    headers = {
        "Content-Type": "application/json",
        "api-key": AZURE_SEARCH_KEY
    }

    keyword_fields = [
        "title", "doc_type", "doc_category", "doc_function",
        "terms", "topics", "summary", "content"
    ]

    metadata_filter_fields = ["doc_type", "doc_category", "doc_function"]
    vector_field = "content_embedding"
    
    optimized_query = llm_search_query_optimizer(query, rewrited_query, use_previous_context)
    keywords = extract_keywords(query, optimized_query, debug=debug_mode) if keywords_matching else None
    query_em = get_query_embedding(optimized_query)

    all_results = []
    all_content = []  # Collect all document content for keyword checking
    index_debug_data = {} 

    for index_name in index_names:
        """Retrieves relevant documents from Azure AI Search"""
        select_fields = [
            "filename", "title", "doc_type", "doc_category", "doc_function",
            "terms", "topics", "summary", "content", "owner", "url"
        ]
        if custom_ranking:
            select_fields.append("content_embedding")

        url = f"{AZURE_SEARCH_ENDPOINT}/indexes/{index_name}/docs/search?api-version=2024-07-01"
        payload = {
            "search": rewrited_query.lower(),
            "count": True,
            "top": top_k,
            "select": ",".join(select_fields),
            "searchFields": ",".join(keyword_fields),    
            "queryType": "semantic",  
            "searchMode": "all",
            "semanticConfiguration": "default-semantic",     
            "vectorQueries": [
                {
                    "kind": "vector",
                    "vector": query_em,
                    "fields": vector_field,
                    "k": top_k*2,
                    "exhaustive": False,
                }
            ]
        }

        if dynamic_filtering and keywords_matching: 
            # Apply dynamic metadata filtering
            generate_field_based_filter(headers, payload, index_name, keywords, filter_fields=metadata_filter_fields)
            if debug:
                print(f"⚙️ Applied filter for index {index_name}: {payload.get('filter', 'None')}")

        response = requests.post(url, headers=headers, json=payload)
        if response.status_code == 200:
            hits = response.json().get("value", [])
            for doc in hits:
                keyword_score = doc.get("@search.rerankerScore") or doc.get("@search.score") or 0.0  
                if custom_ranking:
                    embedding = doc.get("content_embedding")
                    vector_score = cosine_similarity(query_em, embedding) if embedding else 0.0
                    norm_keyword_score = keyword_score / 4.0  # Normalize keyword score to 0-1 range
                    final_score = (1 - vector_weight) * norm_keyword_score + vector_weight * vector_score   
                    doc["_vector_score"] = vector_score
                    doc["_final_score"] = final_score
                else:
                    final_score = keyword_score
                    doc["_final_score"] = final_score

                doc["_index"] = index_name
                doc["_keyword_score"] = keyword_score
                doc["_azure_score"] = doc.get("@search.score", 0.0)
                all_results.append(doc)
                all_content.append(doc.get("content", "").lower() + " " + doc.get("summary", "").lower())

            sorted_hits = sorted(hits, key=lambda x: x.get("_final_score", 0), reverse=True)
            if debug:
                index_debug_data[index_name] = sorted_hits[:3]  # Store top 3
        else:
            logging.warning(f"❌ Search failed on {index_name}: {response.status_code} — {response.text}")
    
    if not all_results:
        return []

    # Compute average top-3 final scores per index
    scores_by_index = {}
    for index in index_names:
        top_docs = sorted([d for d in all_results if d["_index"] == index], key=lambda d: d["_final_score"], reverse=True)[:3]
        if top_docs:
            avg_score = sum(d["_final_score"] for d in top_docs) / len(top_docs)
            scores_by_index[index] = avg_score

    sorted_indexes = sorted(scores_by_index.items(), key=lambda x: x[1], reverse=True)
    best_index, best_score = sorted_indexes[0]

    if debug:
        print("\n🔍 Ranked Indexes by Average Final Score:")
        for name, score in sorted_indexes:
            print(f"  - {name}: {score:.4f}")
        print(f"✅ Best index selected: {best_index} (score: {best_score:.4f})\n")
        
        print("\n🔎 Top 3 chunks from each index:")
        for index, docs in index_debug_data.items():
            print(f"\nIndex: {index}")
            for i, doc in enumerate(docs, 1):
                snippet = doc.get("content", "").replace("\n", " ").strip()
                print(
                    f"  {i}. Score: {doc.get('_final_score', 0):.4f}, "
                    f"Keyword: {doc.get('_keyword_score', 0):.4f}, " +
                    (f"Vector: {doc.get('_vector_score', 0):.4f}, " if custom_ranking else "") +
                    f"Azure: {doc.get('_azure_score', 0):.4f}, " +
                    f"Title: {doc.get('title', 'N/A')}\n     Snippet: {snippet}..."
                )   
    # Return top_k results from best index
    final_results = [doc for doc in all_results if doc["_index"] == best_index]
    final_results = sorted(final_results, key=lambda d: d["_final_score"], reverse=True)[:8]  # Return top 8 as final answer

    warning_msg = ""
    if keywords_matching:
        # --- Check for missing keywords ---
        combined_text = " ".join(all_content)
        missing_keywords = [kw for kw in keywords if kw not in combined_text]

        if missing_keywords:
            warning_msg = f"Keyword search indicates the following words are not found in any relevant documents: '{'; '.join(missing_keywords)}'. You can ignore this message or consider rephrasing your question."
        
    # Final results with warning if needed
    final_answer = warning_msg, final_results
    return final_answer


def multi_index_generate_response(query, context, hide_ref_relevance):
    """
    Generates an answer using OpenAI's o3-mini model with the top chunks, and includes a formatted reference section
    with relevance explanations based on each document's content, summary, topics, and terms.
    """
    client = AzureOpenAI(
        azure_endpoint=AZURE_OPENAI_ENDPOINT,
        api_key=AZURE_OPENAI_API_KEY,
        api_version="2024-12-01-preview"
    )

    warning_msg, top_chunks = context
    context_texts = []
    doc_groups  = {}

    for doc in top_chunks:
        filename = doc.get("filename", "N/A")
        content = doc.get("content", "")
        score = doc.get('_final_score', 0)
        context_texts.append(f"[{filename}] {content}")

        if filename not in doc_groups :
            doc_groups [filename] = {
                "document_name": filename,
                "url": doc.get("url", "N/A"),
                "key_contact": doc.get("owner", "N/A"),
                "key_topics": doc.get("topics", ""),
                "key_terms": doc.get("terms", ""),
                "summary": doc.get("summary", ""),
                "scores": [score],
            }
        else:
            doc_groups [filename]["scores"].append(score)
    
    # --- Build context string for LLM to generate main answer ---
    context_str = "\n\n".join(context_texts)

    # Run LLM-based validation
    is_valid, explanation, is_completely_irrelevant = llm_context_guard_check(query, context_str, client)
    explanation = ' '.join(explanation.strip().split()[1:])   # cleaning the explanation by deleting the first word - yes or no
    if not is_valid:
        # Try to find semi-relevant documents (e.g., score > 0.3) to suggest contact
        main_answer = (
                        "Sorry, I cannot help with that. The provided documents do not clearly explain the requested information.\n"
                        f"(Reason: {explanation})"
        )

        if not is_completely_irrelevant and top_chunks:
            doc = top_chunks[0]
            main_answer += (
                "\n\n---\n Here is relevant document that might be helpful:\n\n"
                "**Related Document**:"
                f"  [{doc.get('filename', 'N/A')}]({doc.get('url', 'N/A')})\n\n"
                f"**Key Contact**: {doc.get('owner', 'N/A')}"
            )
        
        if warning_msg:
            main_answer = f"Notice: {warning_msg.strip()}" + "\n\n" + main_answer
        return f"**Answer:**\n\n{main_answer}"
    
    messages = [
            {
                "role": "user",
                "content": (
                    f"Answer the following question using the context from the top relevant documents:\n"
                    f"QUESTION: {query}\n\n"
                    f"DOCUMENT CHUNKS:\n{context_str}\n\n"
                    f"INSTRUCTIONS:\n- Be concise\n- Use only information from the documents. Do not generate answers that don't use the source documents provided.\n"
                    # f"- If insufficient information or not sure about the answer, ask clarifying questions instead of directly answering it. \n"
                    f"If the answer is not clearly stated in the provided context, or you are unsure, respond with: 'Sorry, I cannot help with that.' Then briely expalain reasoning."
                    f"- You may be provided with multiple sources. Read all sources and find the most relavant information to best answer user question.\n"
                    f"- If your answer describes a process, include step-by-step instructions and seperate each step by bullet symbol.\n"
                    f"- Use plain text with no HTML.\n"
                    f"- Separate sections with line breaks."
                )
            }
        ]

    completion = client.chat.completions.create(
            model=AZURE_OPENAI_DEPLOYMENT,
            messages=messages
    )

    main_answer = completion.choices[0].message.content.strip()
    
    if warning_msg:
            main_answer = f"Notice: {warning_msg.strip()}" + "\n\n" + main_answer

    # --- Build reference section ---
    reference_text = "\n\n**References:**\n"

    for doc in list(doc_groups.values())[:3]:
        if hide_ref_relevance:
            reference_text += (
                f"\n---\n"
                f"**Document**: [{doc['document_name']}]({doc['url']})\n\n"
                f"**Key Contact**: {doc['key_contact']}\n\n"
            )
            continue

        # Build a focused relevance summary prompt
        relevance_context = (
            f"QUESTION: {query}\n\n"
            f"DOCUMENT SUMMARY: {doc['summary']}\n\n"
            f"KEY TOPICS: {doc['key_topics']}\n\n"
            f"KEY TERMS: {doc['key_terms']}\n\n"
        )

        relevance_prompt = [
            {
                "role": "user",
                
                "content": (
                    f"Be concise, Summarize in less than 100 words why this document is relevant to the question below, "
                    f"based only on the document's summary, key topics, and key terms. "
                    f"Use bullet points for clarity.\n"
                    f"Do not include HTML, markdown, or field labels.\n"
                    f"Each bullet point should start with a new line."
                    f"Do NOT list the raw fields; only give a clean, concise explanation.\n\n"
                    f"{relevance_context}"
                )
            }
        ]

        rel_response = client.chat.completions.create(
            model=AZURE_OPENAI_DEPLOYMENT,
            messages=relevance_prompt
        )

        relevance_summary = rel_response.choices[0].message.content.strip()
        avg_score = round(sum(doc["scores"]) / len(doc["scores"]), 2)

        reference_text += (
                f"\n---\n"
                f"**Document**: [{doc['document_name']}]({doc['url']})\n\n"
                f"**Key Contact**: {doc['key_contact']}\n\n"
                # f" **Similarity Score**: {avg_score}\n\n"
                f"**Relevance**: {relevance_summary}\n\n"
        )

    return f"**Answer:**\n\n{main_answer}{reference_text}"


@app.route(route="multiindexquery")
def multiindexquery(req: func.HttpRequest) -> func.HttpResponse:
    logging.info('Python HTTP trigger function processed a request.')
    try:
        req_body = req.get_json()
        query = req_body.get("query")
        query_history = req_body.get("queryhistory", "")
        answer_history = req_body.get("answerhistory", "")
        if not query:
            return func.HttpResponse("Error: Missing 'query' parameter", status_code=400)

        cleaned_query = clean_query_for_llm(query)  # Clean query by removing routing keywords if there are any
        cleaned_query_history = clean_query_for_llm(query_history)

        history_context = filter_relevant_history(cleaned_query, cleaned_query_history, answer_history)
        if use_prev_context:
            rewrited_query = rewrite_query_with_history(cleaned_query, history_context)
        else:
            rewrited_query = cleaned_query

        if metadata_search:
            use_metadata_search_flag = should_use_metadata_search(query)    # use raw query for routing decision
            if use_metadata_search_flag:
                metadata_by_index = metadata_table_by_index(index_names)
                llm_summary = summarize_full_metadata(rewrited_query, history_context, metadata_by_index)        
                return func.HttpResponse(json.dumps({"answer": llm_summary}, ensure_ascii=False, indent=2), mimetype="application/json", status_code=200)

        # Step 1: Search all indexes
        docs = multi_index_search_documents(cleaned_query, rewrited_query, index_names, vector_weight=0.5, top_k=16, 
                                                            dynamic_filtering=dynamic_filtering, 
                                                            keywords_matching = keywords_matching,
                                                            use_previous_context = use_prev_context,    
                                                            custom_ranking=custom_ranking, 
                                                            debug=debug_mode)
        if not docs:
            return func.HttpResponse("No relevant documents found.", status_code=404)
        
        # Step 2: Generate response from AI with retrieved context
        ai_response = multi_index_generate_response(rewrited_query, docs, hide_ref_relevance=hide_ref_relevance)

        return func.HttpResponse(json.dumps({"answer": ai_response}, ensure_ascii=False, indent=2), mimetype="application/json", status_code=200)

    except Exception as e:
        return func.HttpResponse(f"Error: {str(e)}", status_code=500)