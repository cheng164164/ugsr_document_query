import os
import uuid
import base64
import io
import json
import pandas as pd
import requests
from dotenv import load_dotenv
from azure.storage.blob import BlobServiceClient, generate_blob_sas, BlobSasPermissions, BlobClient
from openpyxl import load_workbook
from datetime import datetime, timedelta
from azure.core.credentials import AzureKeyCredential
from azure.ai.documentintelligence import DocumentIntelligenceClient
from azure.ai.documentintelligence.models import AnalyzeDocumentRequest
from azure.search.documents.indexes import SearchIndexClient
from azure.search.documents import SearchClient
from azure.search.documents.indexes.models import (
    SearchIndex, SearchField, SimpleField, SearchableField, VectorSearch,
    VectorSearchProfile, HnswAlgorithmConfiguration, VectorSearchAlgorithmKind,
    SemanticPrioritizedFields, SemanticConfiguration, SemanticField, SemanticSearch, SearchFieldDataType
)
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain_openai import AzureOpenAIEmbeddings
from openai import AzureOpenAI
import re

def read_metadata_from_blob(connection_string, container_name, blob_name):
    blob_service_client = BlobServiceClient.from_connection_string(connection_string)
    blob_client = blob_service_client.get_blob_client(container=container_name, blob=blob_name)

    # Determine file type
    _, file_extension = os.path.splitext(blob_name.lower())
    stream = io.BytesIO()
    blob_data = blob_client.download_blob()
    blob_data.readinto(stream)
    stream.seek(0)

    if file_extension == '.csv':
        df = pd.read_csv(stream)
        df = df.astype(str)
        return df

    elif file_extension in ['.xlsx', '.xls']:
        wb = load_workbook(stream, data_only=True)
        ws = wb.active
        rows = []
        header = [cell.value for cell in ws[1]]
        name_col_index = header.index("Name")
        header.append("url")

        add_url = 'url' not in header
        if add_url:
            header.append("url")

        for row in ws.iter_rows(min_row=2):
            row_values = [cell.value for cell in row]
            if add_url:
                name_cell = row[name_col_index]
                hyperlink = name_cell.hyperlink.target if name_cell.hyperlink else None
                row_values.append(hyperlink)
            rows.append(row_values)

        df = pd.DataFrame(rows, columns=header)
        df = df.astype(str)
        return df
    else:
        raise ValueError("Unsupported file format. Only .xlsx and .csv are supported.")


def create_index(index_name, search_key, search_endpoint):
    credential = AzureKeyCredential(search_key)
    index_client = SearchIndexClient(endpoint=search_endpoint, credential=credential)
    if index_name in [i.name for i in index_client.list_indexes()]:
        index_client.delete_index(index_name)
    fields=[
        SimpleField(name="id", type=SearchFieldDataType.String, key=True, sortable=True, filterable=True, facetable=True),
        SimpleField(name="filename", type=SearchFieldDataType.String),
        SimpleField(name="url", type=SearchFieldDataType.String),
        SearchableField(name="title", type=SearchFieldDataType.String, filterable=True, searchable=True, facetable=True, retrievable=True),
        SearchableField(name="owner", type=SearchFieldDataType.String, filterable=True, searchable=True, facetable=True, retrievable=True),
        SearchableField(name="doc_type", type=SearchFieldDataType.String, filterable=True, searchable=True, facetable=True, retrievable=True),
        SearchableField(name="doc_category", type=SearchFieldDataType.String, filterable=True, searchable=True, facetable=True, retrievable=True),
        SearchableField(name="doc_function", type=SearchFieldDataType.String, filterable=True, searchable=True, facetable=True, retrievable=True),
        SearchableField(name="terms", type=SearchFieldDataType.String, filterable=True, searchable=True, retrievable=True),
        SearchableField(name="topics", type=SearchFieldDataType.String, filterable=True, searchable=True, retrievable=True),
        SearchableField(name="summary", type=SearchFieldDataType.String, filterable=False, facetable=False, searchable=True, retrievable=True),
        SearchableField(name="content", type=SearchFieldDataType.String, filterable=False, facetable=False, searchable=True, retrievable=True),
        SearchField(name="content_embedding", type=SearchFieldDataType.Collection(SearchFieldDataType.Single), searchable=True, hidden=False, vector_search_dimensions=1536, vector_search_profile_name = 'default'),
    ]
    vector_search=VectorSearch(
        algorithms=[
            HnswAlgorithmConfiguration(
                name="default",
                kind=VectorSearchAlgorithmKind.HNSW,
                parameters= {
                    'm': 4,
                    "efConstruction": 400,
                    "efSearch": 500,
                    "metric": "cosine"
                }
            )
        ],
        profiles=[VectorSearchProfile(name="default", algorithm_configuration_name="default")]
    )

    prioritized_fields = SemanticPrioritizedFields(
        title_field= SemanticField(field_name='title'),
        content_fields=[SemanticField(field_name='content'),
                        SemanticField(field_name='summary')],
        keywords_fields=[SemanticField(field_name='terms'),
                        SemanticField(field_name='topics'),
                        SemanticField(field_name='doc_type'),
                        SemanticField(field_name='doc_category'),
                        SemanticField(field_name='doc_function')]
    )

    semantic_configuration = SemanticConfiguration(
        name="default-semantic",
        prioritized_fields=prioritized_fields)
    
    semantic_search = SemanticSearch(
        default_configuration_name="default-semantic",
        configurations=[semantic_configuration]
    )

    index = SearchIndex(name=index_name, fields=fields, semantic_search=semantic_search, vector_search=vector_search)
    index_client.create_index(index)


def create_search_index_from_schema(index_name: str, fields: list, vector_config: dict = None, semantic_config: dict = None):
    """
    Creates or replaces an Azure Search index using the REST API.

    Args:
        index_name (str): Name of the index to create.
        fields (list): List of field definitions (dictionaries).
        vector_config (dict): Optional vector search config (algorithms + profiles).
    """

    search_service = os.getenv("AZURE_SEARCH_ENDPOINT") # e.g., https://your-search.search.windows.net
    search_key = os.getenv("AZURE_SEARCH_KEY")       # Admin key

    if not search_service or not search_key:
        raise ValueError("AZURE_SEARCH_ENDPOINT and AZURE_SEARCH_KEY must be set in environment.")

    headers = {
        "Content-Type": "application/json",
        "api-key": search_key
    }

    index_definition = {
        "name": index_name,
        "fields": fields
    }

    if vector_config:
        index_definition["vectorSearch"] = vector_config
    if semantic_config:
        index_definition["semantic"] = semantic_config

    url = f"{search_service}/indexes/{index_name}?api-version=2023-10-01-Preview"
    response = requests.put(url, headers=headers, json=index_definition)

    if response.status_code in [200, 201]:
        print(f"✅ Index '{index_name}' created or updated successfully.")
    else:
        print(f"❌ Failed to create index: {response.status_code}")
        print(response.text)


def list_blobs(connection_string, container_name):
    blob_service_client = BlobServiceClient.from_connection_string(connection_string)
    container_client = blob_service_client.get_container_client(container_name)
    return [b for b in container_client.list_blobs() if b.name != "index_log.csv"]

def generate_blob_sas_url(connection_string, container_name, blob_name):
    blob_service_client = BlobServiceClient.from_connection_string(connection_string)
    blob_client = blob_service_client.get_blob_client(container=container_name, blob=blob_name)
    sas_token = generate_blob_sas(
        account_name=blob_client.account_name,
        container_name=container_name,
        blob_name=blob_name,
        account_key=blob_service_client.credential.account_key,
        permission=BlobSasPermissions(read=True),
        expiry=datetime.utcnow() + timedelta(hours=1)
    )
    return f"{blob_client.url}?{sas_token}"

def document_read(sas_url, azure_doc_intell_endpoint, azure_doc_intell_key):
    client = DocumentIntelligenceClient(endpoint=azure_doc_intell_endpoint, credential=AzureKeyCredential(azure_doc_intell_key))
    poller = client.begin_analyze_document("prebuilt-read", AnalyzeDocumentRequest(url_source=sas_url))
    result = poller.result()
    return result.content

def obtain_topics(context, azure_oai_endpoint, azure_oai_key, azure_oai_deployment_model):
    client = AzureOpenAI(
        azure_endpoint=azure_oai_endpoint,
        api_key=azure_oai_key,
        api_version="2025-01-01-preview",
    )
    messages = [{"role": "user", "content": f"Obtain all the main topics mentioned in this document. Keep your response short and just include the topics. Also, include the scope and purpose. Here is the document Content: {context}"}]
    completion = client.chat.completions.create(
        model=azure_oai_deployment_model,
        messages=messages
    )
    return completion.choices[0].message.content

def obtain_key_terms(context, azure_oai_endpoint, azure_oai_key, azure_oai_deployment_model):
    client = AzureOpenAI(
        azure_endpoint=azure_oai_endpoint,
        api_key=azure_oai_key,
        api_version="2025-01-01-preview",
    )
    messages = [{"role": "user", "content": f"Obtain key terminology used in this document. Keep your response short and just include the terms. Here is the document Content: {context}"}]
    completion = client.chat.completions.create(
        model=azure_oai_deployment_model,
        messages=messages
    )
    return completion.choices[0].message.content

def obtain_summary(context, azure_oai_endpoint, azure_oai_key, azure_oai_deployment_model):
    client = AzureOpenAI(
        azure_endpoint=azure_oai_endpoint,
        api_key=azure_oai_key,
        api_version="2025-01-01-preview",
    )
    messages = [{"role": "user", "content": f"Obtain the summary of the document. Keep your response short under 200 words. Here is the document Content: {context}"}]
    completion = client.chat.completions.create(
        model=azure_oai_deployment_model,
        messages=messages
    )
    return completion.choices[0].message.content

def truncate_summary(text, max_chars=4000):
    if len(text) <= max_chars:
        return text
    truncated = text[:max_chars]
    end = truncated.rfind(". ")
    return truncated[:end+1] if end > 0 else truncated


def obtain_version_and_publish_date(context, azure_oai_endpoint, azure_oai_key, azure_oai_deployment_model):
    from openai import AzureOpenAI

    client = AzureOpenAI(
        azure_endpoint=azure_oai_endpoint,
        api_key=azure_oai_key,
        api_version="2025-01-01-preview",
    )

    messages = [
        {
            "role": "user",
            "content": (
                "From the following document content, extract the **version number** and **publish date** "
                "(or effective date) if they exist. These are usually written together in a field like: \"Ver. 1.0, 05-25\"."
                "or the date maybe in format like 12/4/2024, the version maybe in format like 3.0. Please find the latest date and version.\n"
                "Return only in the following JSON format:\n"
                "{\"version\": <version>, \"publish_date\": <date>}\n"
                "If a field is not found, set it to null. Here is the document content: " + context[:4000]
            )
        }
    ]

    try:
        completion = client.chat.completions.create(
            model=azure_oai_deployment_model,
            messages=messages
        )
        result = completion.choices[0].message.content.strip()
        parsed = json.loads(result)
        return parsed.get("version"), parsed.get("publish_date")
    except Exception as e:
        return None, None


def save_metadata_to_blob(metadata_df, connection_string, container_name, blob_name):
    output_csv = metadata_df.to_csv(index=False)
    blob_client = BlobClient.from_connection_string(
        conn_str=connection_string,
        container_name=container_name,
        blob_name=blob_name
    )
    blob_client.upload_blob(output_csv, overwrite=True)


def is_english_filename(name):
    try:
        # Allow most common characters in English filenames
        return bool(re.match(r"^[A-Za-z0-9\s\-\u2013\u2014\.,_()\[\]{}'\":;@&!#\$%\^+=]+$", name))
    except:
        return False


def make_doc_id(file_key: str, chunk_id: int) -> str:
    """Return a stable document ID for Azure AI Search."""
    encoded = base64.urlsafe_b64encode(file_key.encode("utf-8")).decode("utf-8").rstrip("=")
    return f"{encoded}-chunk-{chunk_id}"


def chunk_and_embed_docs(splitter, embedder, embedder_client, connection_string, container_name, metadata_df, metadata_container, metadata_blob_name,
                         azure_doc_intell_endpoint, azure_doc_intell_key, azure_oai_endpoint, azure_oai_key, 
                         azure_oai_deployment_model, using_embedder=True):
    
    # Ensure version and publish_date columns exist
    if "version" not in metadata_df.columns:
        metadata_df["version"] = None
    if "publish date" not in metadata_df.columns:
        metadata_df["publish date"] = None
    
    indexed_docs = []
    blob_list = list_blobs(connection_string, container_name)
    for blob in blob_list:
        # Skip non-English file names
        if not is_english_filename(blob.name):
            print(f"⏭️ Skipped non-English file: {blob.name}")
            continue
        
        file_name = blob.name
        sas_url = generate_blob_sas_url(connection_string, container_name, file_name)
        doc_content = document_read(sas_url, azure_doc_intell_endpoint, azure_doc_intell_key)
        topics = obtain_topics(doc_content, azure_oai_endpoint, azure_oai_key, azure_oai_deployment_model)
        terms = obtain_key_terms(doc_content, azure_oai_endpoint, azure_oai_key, azure_oai_deployment_model)
        summary = truncate_summary(obtain_summary(doc_content, azure_oai_endpoint, azure_oai_key, azure_oai_deployment_model))
        ver, date = obtain_version_and_publish_date(doc_content, azure_oai_endpoint, azure_oai_key, azure_oai_deployment_model)
        meta_row = metadata_df[metadata_df["Name"].str.lower() == os.path.basename(file_name).lower()]
        if meta_row.empty:
            continue
        idx = meta_row.index[0]
        metadata_df.at[idx, "version"] = ver
        metadata_df.at[idx, "publish date"] = date

        meta = meta_row.iloc[0].to_dict()
        chunks = splitter.create_documents([doc_content])
        if using_embedder:
            vectors = embedder.embed_documents([chunk.page_content for chunk in chunks])
            # vector_summary = embedder.embed_documents([summary])[0]
        else:
            vectors = embedder_client.embeddings.create(model="text-embedding-3-small", input=[chunk.page_content for chunk in chunks])
            # vector_summary = embedder_client.embeddings.create(model="text-embedding-3-small", input=[summary])

        for chunk_id, (vec, chunk) in enumerate(zip(vectors, chunks)):
            indexed_docs.append({
                "id": make_doc_id(file_name.lower(), chunk_id),
                "filename": file_name.lower(),
                "title": meta.get("Title", "").strip().lower(),
                "url": meta["url"],
                "owner": meta.get("Document Owner(s)", "").strip().lower(),
                "doc_type": meta.get("Doc Type", "").strip().lower(),
                "doc_category": meta.get("Doc Category", "").strip().lower(),
                "doc_function": meta.get("Function", "").strip().lower(),
                "terms": terms,
                "topics": topics,
                "summary": summary,
                "content": chunk.page_content,
                "content_embedding": vec,
            })

    # Save updated metadata_df to blob using separate function
    save_metadata_to_blob(metadata_df, connection_string, container_name, metadata_blob_name)
    return indexed_docs


def upload_search_index(index_name, search_key, search_endpoint, indexed_docs):
    credential = AzureKeyCredential(search_key)
    search_client = SearchClient(endpoint=search_endpoint, index_name=index_name, credential=credential)
    for i in range(0, len(indexed_docs), 1000):
        search_client.upload_documents(documents=indexed_docs[i:i+1000])


def data_chunk_embed_upload_batch(splitter, embedder, embedder_client, connection_string, container_name, metadata_df, metadata_container, metadata_blob_name,
                  index_name, azure_doc_intell_endpoint, azure_doc_intell_key, azure_oai_endpoint, 
                  azure_oai_key, azure_oai_deployment_model, using_embedder=True, batch_number=0, batch_size=30, total_batches=None, blob_subset=None):
    '''
    Manually control batch run process by assigning batch_number.
    '''
    from azure.storage.blob import ContainerClient
    from azure.core.credentials import AzureKeyCredential
    from azure.search.documents import SearchClient

    # Ensure version and publish_date columns exist
    if "version" not in metadata_df.columns:
        metadata_df["version"] = None
    if "publish date" not in metadata_df.columns:
        metadata_df["publish date"] = None

    # --- Load blob list
    container_client = ContainerClient.from_connection_string(connection_string, container_name)
    all_blobs = list(container_client.list_blobs())
    if blob_subset is not None:
        # Use passed subset directly
        blob_subset_set = set(blob_subset)  # For faster lookup
        current_batch = [b for b in all_blobs if b.name in blob_subset_set]
        total_files = len(current_batch)
        start = 0
        end = total_files
    else:
        # Use calculated slicing based on batch_number and batch_size
        blob_list = [b for b in all_blobs if b.name != "index_log.csv"]
        total_files = len(blob_list)
        start = batch_number * batch_size
        end = min(start + batch_size, total_files)
        current_batch = blob_list[start:end]

    if not current_batch:
        print(f"❌ No files found in batch {batch_number}")
        return

    print(f"\n📦 [Index: {index_name}] Starting batch {batch_number + 1}: processing files {start + 1} to {end} of {total_files}")
    indexed_docs = []

    for i, blob in enumerate(current_batch):
        print(f"📄 [Index: {index_name} ({total_files} files)] [Batch {batch_number + 1}/{total_batches}] [{start + i + 1}/{total_files}] Processing: {blob.name}")

        # Skip non-English file names
        if not is_english_filename(blob.name):
            print(f"⏭️ Skipped non-English file: {blob.name}")
            continue
        
        try:
            file_name = blob.name
            sas_url = generate_blob_sas_url(connection_string, container_name, blob.name)
            doc_content = document_read(sas_url, azure_doc_intell_endpoint, azure_doc_intell_key)
            topics = obtain_topics(doc_content, azure_oai_endpoint, azure_oai_key, azure_oai_deployment_model)
            terms = obtain_key_terms(doc_content, azure_oai_endpoint, azure_oai_key, azure_oai_deployment_model)
            summary = truncate_summary(obtain_summary(doc_content, azure_oai_endpoint, azure_oai_key, azure_oai_deployment_model))
            ver, date = obtain_version_and_publish_date(doc_content, azure_oai_endpoint, azure_oai_key, azure_oai_deployment_model)

            # print(f"doc_content: {doc_content}\n** topics: {topics}\n** terms: {terms}\n** summary: {summary}")
            meta_row = metadata_df[metadata_df["Name"].str.lower() == os.path.basename(blob.name).lower()]
            if meta_row.empty:
                print(f"⚠️  Skipped (no metadata): {blob.name}")
                continue

            idx = meta_row.index[0]
            metadata_df.at[idx, "version"] = ver
            metadata_df.at[idx, "publish date"] = date

            meta = meta_row.iloc[0].to_dict()
            chunks = splitter.create_documents([doc_content])

            if using_embedder:
                vectors = embedder.embed_documents([chunk.page_content for chunk in chunks])
                # vector_summary = embedder.embed_documents([summary])[0]
            else:
                vectors = embedder_client.embeddings.create(model="text-embedding-3-small", input=[chunk.page_content for chunk in chunks])
                # vector_summary = embedder_client.embeddings.create(model="text-embedding-3-small", input=[summary])

            for chunk_id, (vec, chunk) in enumerate(zip(vectors, chunks)):
                indexed_docs.append({
                    "id": make_doc_id(file_name.lower(), chunk_id),
                    "filename": file_name.lower(),
                    "title": meta.get("Title", "").strip().lower(),
                    "url": meta["url"],
                    "owner": meta.get("Document Owner(s)", "").strip().lower(),
                    "doc_type": meta.get("Doc Type", "").strip().lower(),
                    "doc_category": meta.get("Doc Category", "").strip().lower(),
                    "doc_function": meta.get("Function", "").strip().lower(),
                    "terms": terms,
                    "topics": topics,
                    "summary": summary,
                    "content": chunk.page_content,
                    "content_embedding": vec,
                    # "summary_embedding": vector_summary
                })
        
        except Exception as e:
            print(f" ❌ Failed: {blob.name} — {str(e)}")
            continue

    # Save updated metadata_df to blob using separate function
    save_metadata_to_blob(metadata_df, connection_string, metadata_container, metadata_blob_name)

    # --- Upload to Azure Search
    print('Now uploading the indexed_docs to the index created in AI search service...')
    credential = AzureKeyCredential(os.getenv("AZURE_SEARCH_KEY"))
    search_client = SearchClient(endpoint=os.getenv("AZURE_SEARCH_ENDPOINT"), index_name=index_name, credential=credential)

    for i in range(0, len(indexed_docs), 1000):
        result = search_client.upload_documents(documents=indexed_docs[i:i+1000])
        for r in result:
            if not r.succeeded:
                print(f"❌ [Index: {index_name}] [Batch {batch_number + 1}/{total_batches or '?'}] Upload failed: {r.key} — {r.error_message}")
        print(f" ✅ [Index: {index_name}] [Batch {batch_number + 1}/{total_batches or '?'}] Uploaded batch segment {i//1000 + 1}"
              f"({min(i+1000, len(indexed_docs))}/{len(indexed_docs)} chunks)")

    print(f" 🎉 Finished uploading [Index: {index_name}] batch {batch_number + 1} ({len(indexed_docs)} chunks)")

    return metadata_df
